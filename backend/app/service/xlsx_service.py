from pathlib import Path
from typing import Any

import tiktoken
from openpyxl import load_workbook

from app.config.settings import get_settings
from app.service.docmind_service import DocMindService


ROWS_PER_CHUNK = 30
MAX_EXCEL_CHUNK_TOKENS = 800
DOCMIND_XLSX_MAX_BYTES = 5 * 1024 * 1024
MAX_COLUMNS_PER_ROW = 24
MAX_EXCEL_CHUNKS = 1000
IDENTIFIER_COLUMN_NAMES = {"company_name", "code", "industry", "statement_type", "date"}
PRIORITY_COLUMN_KEYWORDS = (
    "net_income",
    "total_revenue",
    "operating_revenue",
    "gross_profit",
    "operating_income",
    "total_assets",
    "total_liab",
    "stockholders_equity",
    "cash",
    "eps",
)
IMPORTANT_COLUMN_KEYWORDS = (
    "revenue",
    "income",
    "profit",
    "net_income",
    "assets",
    "liabil",
    "equity",
    "cash",
    "eps",
    "ebit",
    "expense",
    "debt",
    "operating",
)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _encoding():
    settings = get_settings()
    try:
        return tiktoken.encoding_for_model(settings.embedding_model)
    except KeyError:
        return tiktoken.get_encoding("cl100k_base")


def _split_large_chunk(
    chunk: dict[str, Any],
    max_tokens: int = MAX_EXCEL_CHUNK_TOKENS,
) -> list[dict[str, Any]]:
    content = str(chunk["content"]).strip()
    if not content:
        return []
    encoding = _encoding()
    token_ids = encoding.encode(content)
    if len(token_ids) <= max_tokens:
        return [chunk]

    chunks: list[dict[str, Any]] = []
    for part_index, start in enumerate(range(0, len(token_ids), max_tokens)):
        part = encoding.decode(token_ids[start : start + max_tokens]).strip()
        if not part:
            continue
        split_chunk = dict(chunk)
        split_chunk["content"] = part
        split_chunk["chunk_part"] = part_index
        chunks.append(split_chunk)
    return chunks


def _limit_chunks(chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return chunks[:MAX_EXCEL_CHUNKS]


def parse_xlsx_with_docmind(file_path: str, file_name: str) -> str:
    service = DocMindService()
    task_id = service.submit_job(file_path, file_name)
    if not task_id:
        raise RuntimeError("DocMind Excel 任务提交失败")
    if not service.wait_for_completion(task_id):
        raise RuntimeError("DocMind Excel 解析超时或失败")
    text = service.collect_all_results(task_id)
    if not text.strip():
        raise RuntimeError("DocMind Excel 解析结果为空")
    return text


def _select_column_indexes(header: list[str]) -> list[int]:
    if len(header) <= MAX_COLUMNS_PER_ROW:
        return list(range(len(header)))

    selected: list[int] = []
    for index, name in enumerate(header):
        normalized = name.strip().lower()
        if normalized in IDENTIFIER_COLUMN_NAMES:
            selected.append(index)

    for index, name in enumerate(header):
        if index in selected:
            continue
        normalized = name.strip().lower()
        if any(keyword in normalized for keyword in PRIORITY_COLUMN_KEYWORDS):
            selected.append(index)
        if len(selected) >= MAX_COLUMNS_PER_ROW:
            break

    for index, name in enumerate(header):
        if index in selected:
            continue
        normalized = name.strip().lower()
        if any(keyword in normalized for keyword in IMPORTANT_COLUMN_KEYWORDS):
            selected.append(index)
        if len(selected) >= MAX_COLUMNS_PER_ROW:
            break

    if not selected:
        selected = list(range(min(len(header), MAX_COLUMNS_PER_ROW)))
    return sorted(selected[:MAX_COLUMNS_PER_ROW])


def _format_row(
    row: list[str],
    header: list[str],
    selected_indexes: list[int],
    compact_mode: bool,
) -> str:
    values: list[str] = []
    for index in selected_indexes:
        value = row[index] if index < len(row) else ""
        if not value:
            continue
        if compact_mode:
            column_name = header[index] if index < len(header) and header[index] else f"col_{index + 1}"
            values.append(f"{column_name}={value}")
        else:
            values.append(value)
    return " | ".join(values)


def parse_xlsx(file_path: str) -> list[dict[str, Any]]:
    """Fallback XLSX parser using openpyxl."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [
                [_cell_to_text(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            non_empty_rows = [row for row in rows if any(row)]
            if not non_empty_rows:
                continue

            header = non_empty_rows[0]
            selected_indexes = _select_column_indexes(header)
            compact_mode = selected_indexes != list(range(len(header)))
            data_rows = non_empty_rows[1:] or [header]
            for start in range(0, len(data_rows), ROWS_PER_CHUNK):
                row_batch = data_rows[start : start + ROWS_PER_CHUNK]
                lines = [f"# Sheet: {worksheet.title}"]
                if data_rows is not non_empty_rows:
                    lines.append(
                        "表头: " + " | ".join(header[index] for index in selected_indexes)
                    )
                for offset, row in enumerate(row_batch, start=start + 2):
                    values = _format_row(row, header, selected_indexes, compact_mode)
                    if not values:
                        continue
                    lines.append(f"Row {offset}: {values}")
                chunks.extend(
                    _split_large_chunk(
                        {
                            "content": "\n".join(lines).strip(),
                            "source_type": "xlsx",
                            "sheet_name": worksheet.title,
                            "row_start": start + 2,
                            "row_end": start + len(row_batch) + 1,
                        }
                    )
                )
    finally:
        workbook.close()
    return _limit_chunks(chunks)


def chunk_excel_text(
    text: str,
    *,
    source_type: str = "excel",
    rows_per_chunk: int = ROWS_PER_CHUNK,
) -> list[dict[str, Any]]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    chunks: list[dict[str, Any]] = []
    for start in range(0, len(lines), rows_per_chunk):
        batch = lines[start : start + rows_per_chunk]
        chunks.extend(
            _split_large_chunk(
                {
                    "content": "\n".join(batch),
                    "source_type": source_type,
                    "row_start": start + 1,
                    "row_end": start + len(batch),
                }
            )
        )
    return _limit_chunks(chunks)


def parse_xlsx_to_chunks(file_path: str, file_name: str) -> list[dict[str, Any]]:
    if Path(file_path).stat().st_size > DOCMIND_XLSX_MAX_BYTES:
        return parse_xlsx(file_path)

    try:
        text = parse_xlsx_with_docmind(file_path, file_name)
        chunks = chunk_excel_text(text, source_type="xlsx_docmind")
        if chunks:
            return chunks
    except Exception:
        pass
    return parse_xlsx(file_path)


def parse_xls_to_chunks(file_path: str, file_name: str) -> list[dict[str, Any]]:
    text = parse_xlsx_with_docmind(file_path, file_name)
    chunks = chunk_excel_text(text, source_type=Path(file_name).suffix.lstrip(".") or "xls")
    if not chunks:
        raise RuntimeError("Excel 解析结果为空")
    return chunks
